#!python

import os
import win32com.client as win32
from openpyxl import load_workbook
from argparse import ArgumentParser

BASEPATH = 'Z:\\new-app'

def print_file(filepath):
    wob = load_workbook(filepath)
    fullpath = BASEPATH + '\\' + filepath

    # wake up the excel application
    excel_app = win32.gencache.EnsureDispatch('Excel.Application')

    # make the client visible
    excel_app.Visible = True
    print_wb = excel_app.Workbooks.Open(fullpath)

    # select all sheets
    print_worksheets = print_wb.Worksheets(wob.sheetnames)

    # send print job
    print_worksheets.PrintOut()

    # close and quit
    print_wb.Close(False)
    wob.close()
    excel_app.Quit()

if __name__ == '__main__':
    parser = ArgumentParser(description='Print an Excel file.')
    parser.add_argument('filepath', type=str, help='The path to the Excel file to print.')
    args = parser.parse_args()

    print_file(args.filepath)
